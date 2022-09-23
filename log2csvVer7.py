from pickle import APPEND
from pyexpat import model
import sys
import pandas as pd
import openpyxl
def prcsLine(line):
    print(line)

def atoi(s):
    s = s[::-1]
    num = 0
    for i, v in enumerate(s):
        for j in range(0, 10):
            if v == str(j):
                num += j * (10 ** i)
    return num

data1 = {"curPos":[],"curSpd":[],"targSpd":[],"pid":[]}
list0=[]
list1=[]
list2=[]
list3=[]

def line2df(s):
    list = s.split(',')
    list[0]= int(list[0]) 
    list[1]= int(list[1]) 
    list[2]= int(list[2]) 
    list[3]= int(list[3])
    list0.append(list[0])
    list1.append(list[1])
    list2.append(list[2])
    list3.append(list[3])
    data1["curPos"]=list0
    data1["curSpd"]=list1
    data1["targSpd"]=list2
    data1["pid"]=list3
    return data1

def clearData():
    global data1 
    data1 = {"curPos":[],"curSpd":[],"targSpd":[],"pid":[]}
    global list0
    list0=[]
    global list1
    list1=[]
    global list2
    list2=[]
    global list3
    list3=[]    

d1 = {"curPos":[],"curSpd":[],"targSpd":[],"pid":[]}
l0=[]
l1=[]
l2=[]
l3=[]
def line2df2(s):
    list = s.split(',')
    list[0]= int(list[0]) 
    list[1]= int(list[1]) 
    list[2]= int(list[2]) 
    list[3]= int(list[3])
    l0.append(list[0])
    l1.append(list[1])
    l2.append(list[2])
    l3.append(list[3])
    d1["curPos"]=list0
    d1["curSpd"]=list1
    d1["targSpd"]=list2
    d1["pid"]=list3
    return d1

def clearShtData():
    global d1 
    d1 = {"curPos":[],"curSpd":[],"targSpd":[],"pid":[]}
    global l0
    l0=[]
    global l1
    l1=[]
    global l2
    l2=[]
    global l3
    l3=[]

def line2list(s):
    list = s.split(',')
    list[0]= int(list[0]) 
    list[1]= int(list[1]) 
    list[2]= int(list[2]) 
    list[3]= int(list[3])
    return list

def main():
    filename = sys.argv[1]
#    fix1=[1,9 ,8 ,7, 6, 5 ,4, 3, 2, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1,  1,  1,  1]
#    fix2=[1,10,10,10,10,10,10,10,10,10,15,20,30,40,60,70,80,90,100,200,400,800]
    fix1=[1,1,1,1,1,1,1,1,1]
    fix2=[1000,3000,5000,7000,10000,30000,50000,70000,100000]

#    filename1= filename+".csv"
#    fc = open(filename1, "w")
#    fc.writelines("curPos,  curSpd,  targSpd, pid"+"\n")
    start = 0
    booknum = 0
    with open(filename,'r') as f:
        line =f.readline()
        
#        book = openpyxl.Workbook()
#        sheetmain = book.active
#        sheet1 = book.create_sheet()
#        sheetname = 0
#        linei = 1
#        linej = 1
#        name =""

        while line:
#            print(line)
            if(line.find("jisstart") != -1):
                clearData()
                if(booknum>0):
                    book.save(name) 
                booknum = booknum+1
                index = line.find("j = [")
                name = line[index+5:-2]
                name = int(name)
                name = str(fix1[name])+"%"+str(fix2[name])
                name = name + ".xlsx"
                book = openpyxl.Workbook()
                sheetmain = book.active
                sheetmain.title = name
                sheetmain.cell(row=1,column=1,value="curPos")
                sheetmain.cell(row=1,column=2,value="curSpd")
                sheetmain.cell(row=1,column=3,value="targSpd")
                sheetmain.cell(row=1,column=4,value="pid")
                sheetname = 0
                linei=2
#                fc = open(name, "w")
#                fc.writelines("curPos,  curSpd,  targSpd, pid"+"\n")

            if(line.find("state  curPos  curSpd  targSpd pid") != -1):
                sheet1 = book.create_sheet(str(sheetname))
                sheet1.cell(row=1,column=1,value="curPos")
                sheet1.cell(row=1,column=2,value="curSpd")
                sheet1.cell(row=1,column=3,value="targSpd")
                sheet1.cell(row=1,column=4,value="pid")
                
                sheetname=sheetname+1
#                clearShtData()
                print(sheetname)
                start = 1
                linej = 2
            if(start == 1) and (line.find("M31:ENG")!=-1):
                if((line.find("8 ,") != -1) or (line.find("64 ,") != -1)or (line.find("1 ,") != -1)):
#                    print(line)
                    index = line.find("1 ,")
                    if(index != -1):
                        line1 = line[index+3:-1]
                        df = line2df(line1)
                        shtdf = line2df2(line1)
                        list1 = line2list(line1)
#                        fc.writelines(line1+"\n")
                    else:
                        index = line.find("8 ,")
                        if(index != -1):
                            line1 = line[index+3:-1]
                            df = line2df(line1)
                            shtdf = line2df2(line1)
                            list1 = line2list(line1)
#                            fc.writelines(line1+"\n")
                        else:
                            index = line.find("64 ,")
                            if(index != -1):
                                line1 = line[index+4:-1]
                                df = line2df(line1)
                                shtdf = line2df2(line1)
                                list1 = line2list(line1)
#                                fc.writelines(line1+"\n")
                    #把list内容加入到sheet里面
                    sheetmain.cell(row=linei,column=1,value=list1[0])
                    sheetmain.cell(row=linei,column=2,value=list1[1])
                    sheetmain.cell(row=linei,column=3,value=list1[2])
                    sheetmain.cell(row=linei,column=4,value=list1[3])
                    sheet1.cell(row=linej,column=1,value=list1[0])
                    sheet1.cell(row=linej,column=2,value=list1[1])
                    sheet1.cell(row=linej,column=3,value=list1[2])
                    sheet1.cell(row=linej,column=4,value=list1[3])
                    linei=linei+1
                    linej=linej+1
            line=f.readline()
        book.save(name)
    f.close()
    
main()